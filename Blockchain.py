import hashlib
import time
import random
from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from Crypto.PublicKey import RSA
from Crypto.Signature import PKCS1_v1_5
from Crypto.Hash import SHA256

from Transacao import Transacao
from Bloco import Bloco 

class Blockchain:
    def __init__(self, tamanho_nonce=100):
        self.cadeia = [self.criar_bloco_genese()]
        self.dificuldade = 4
        self.tamanho_nonce = tamanho_nonce  # Inicializando o atributo tamanho_nonce
        self.transacoes_pendentes = []
        self.nos = []
        self.blocos_descartados = []
        self.hash_total = 0
        self.tempo_processamento = 0
        self.blocos_criados = 0

    def criar_bloco_genese(self):
        return Bloco([], "0")

    def adicionar_bloco(self, novo_bloco):
        if len(self.cadeia) == 1:
            print("Iniciando a mineração dos blocos para criação da Blockchain.")
        novo_bloco.hash_anterior = self.cadeia[-1].hash
        inicio = time.time()
        novo_bloco.minerar_bloco(self.dificuldade)
        fim = time.time()
        self.tempo_processamento += fim - inicio
        if novo_bloco.proof is not None:
            self.cadeia.append(novo_bloco)
            self.blocos_criados += 1
        else:
            self.blocos_descartados.append(novo_bloco)

    def get_dificuldade(self):
        return self.dificuldade

    def get_tamanho_nonce(self):
        return self.tamanho_nonce

    def adicionar_transacao(self, transacao):
        self.transacoes_pendentes.append(transacao)

    def processar_transacoes_pendentes(self, no_minerador):
        novo_bloco = Bloco(self.transacoes_pendentes, self.cadeia[-1].hash)
        self.adicionar_bloco(novo_bloco)
        self.transacoes_pendentes = []
        for no in self.nos:
            if no != no_minerador:
                no.atualizar_cadeia(self.cadeia)

    def calcular_taxa_hash(self):
        tempo_inicial = time.time()
        contador = 0
        while time.time() < tempo_inicial + 60:  # Calcula a taxa de hash por minuto
            contador += 1
            Bloco([], "0").minerar_bloco(self.dificuldade)
        return contador

    def run_simulacao(self, total_blocos):
        inicio = time.time()
        for _ in range(total_blocos):  # Definir a quantidade de blocos conforme a escolha do usuário
            novo_bloco = Bloco(self.transacoes_pendentes, self.cadeia[-1].hash)
            self.adicionar_bloco(novo_bloco)
            self.transacoes_pendentes = []
            time.sleep(0.001)  # Apenas para simular o processo em tempo real
        fim = time.time()

        # Exibindo mensagem de conclusão da mineração e criação da blockchain
        print("\nExibindo estatística da Blockchain processada.")

        # Criando a tabela com as estatísticas
        tabela = PrettyTable()
        tabela.field_names = ["Estatística", "Valor"]
        tabela.add_row(["Blocos criados", self.blocos_criados])
        tabela.add_row(["Blocos válidos", len(self.cadeia) - len(self.blocos_descartados)])
        tabela.add_row(["Blocos descartados", len(self.blocos_descartados)])
        eficiencia = (len(self.cadeia) - len(self.blocos_descartados)) / len(self.cadeia) * 100
        taxa_hash_total = self.calcular_taxa_hash()  # Calcula a taxa de hash por minuto
        tabela.add_row(["Taxa de hash por minuto total", taxa_hash_total])

        # Exibindo a tabela
        print(tabela)

        # Salvar estatísticas em um arquivo de texto
        with open("estatisticas_blockchain.txt", "w") as file:
            file.write(str(tabela))

        # Exportando blocos para planilha Excel
        wb = Workbook()
        ws = wb.active
        ws.append(["Bloco", "Hash", "Hash Anterior", "Timestamp", "Transações"])
        for i, bloco in enumerate(self.cadeia):
            cor = '00FF00' if bloco not in self.blocos_descartados else 'FF0000'
            ws.append([f"Bloco {i+1}", bloco.hash, bloco.hash_anterior, bloco.timestamp, str(bloco.transacoes)])
            ws[f"A{i+2}"].fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        wb.save("process_blockchain_tcc_geovane.xlsx")

        # Exibindo mensagem de conclusão da mineração e criação da blockchain
        print("\nMineração concluída. Blockchain criada.")
        input("\nPressione Enter para continuar.")  # Aguardando o pressionamento de Enter para continuar

if __name__ == "__main__":
    # Solicitar ao usuário o número de blocos a serem criados
    try:
        total_blocos = int(input("Digite o número de blocos a serem criados na blockchain: "))
    except ValueError:
        print("Por favor, insira um número válido.")
        exit()

    blockchain = Blockchain()
    blockchain.run_simulacao(total_blocos)